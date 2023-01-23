from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook("merged_data.xlsx")

# Count the number of worksheets
worksheet_count = len(workbook.worksheets)
print("Number of worksheets:", worksheet_count)


'''
This code is a python script that uses the openpyxl library to load an existing excel file named "merged_data.xlsx". It then uses the "load_workbook()" function to load the workbook into the variable "workbook". The script then uses the "len()" function to count the number of worksheets in the workbook and assigns the value to the variable "worksheet_count". Finally, it uses the "print()" function to display the message "Number of worksheets: " followed by the value of "worksheet_count" on the console. This code is used to count the number of sheets in an excel file.
'''
