#!usr/bin/python3
# Import the openpyxl library
from openpyxl import Workbook, load_workbook

# List of file names
file_list = ["Persona.xlsx", "next.xlsx"]

# Load the Persona workbook as the primary file
merged_book = load_workbook(file_list[0])

# Iterate through the list of file names starting from the second file
for file in file_list[1:]:
    # Load the workbook
    book = load_workbook(file)
    # Iterate over the sheets
    for sheet in book:
        # Copy the sheet to the merged workbook
        merged_book.create_sheet(title=sheet.title, index=len(merged_book.sheetnames))
        # Copy the data to the new sheet
        for row in sheet.iter_rows():
            for cell in row:
                merged_book[sheet.title][cell.coordinate].value = cell.value

# Save the merged workbook
merged_book.save("merged_data.xlsx")
